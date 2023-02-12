import json
import os
import tempfile
import unittest
from unittest.mock import patch
from openpyxl import load_workbook, styles

from unexpected_isaves.save_image import to_excel, to_minecraft, to_ascii

IMG_PATH = "./tests/assets/python-logo.png"


class TestToExcel(unittest.TestCase):
    def test_default(self):
        outfile_path = tempfile.mkstemp()[1] + ".xlsx"
        try:
            to_excel(image=IMG_PATH, path=outfile_path)
            wb = load_workbook(outfile_path)
        finally:
            os.remove(outfile_path)
        ws = wb.active
        # decode expected value from json file
        __location__ = os.path.realpath(
            os.path.join(os.getcwd(), os.path.dirname(__file__))
        )
        expected_file = open(
            os.path.join(__location__, "fixtures/to_excel_expected.json"), "r"
        )
        expected_default = json.load(expected_file)
        expected_file.close()
        # verifies one cell, one row further than the image
        for r in range(0, 204 + 1):
            for c in range(0, 186 + 1):
                cell = ws.cell(row=r + 1, column=c + 1)
                h = cell.fill.start_color.index
                e = expected_default[r][c]
                m1 = f"Cell {r+1}:{c+1} should be {e}, received {h}"
                self.assertEqual(h, e, m1)
                m2 = f"Cell {r+1}:{c+1} should be empty"
                self.assertEqual(cell.value, None, m2)


class TestToMinecraft(unittest.TestCase):
    @patch("unexpected_isaves.save_image.__to_minecraft_save")
    def test_default(self, mock_to_minecraft_save):
        to_minecraft(image=IMG_PATH, path="mustnt_save", lower_image_size_by=50)
        # asset __to_minecraft_save was called
        mock_to_minecraft_save.assert_called_once()
        mock_to_minecraft_save.assert_called_with(
            [
                "fill 0 0 0 11 0 0 minecraft:black_wool",
                "fill 12 0 0 12 0 0 minecraft:b",
                "fill 13 0 0 13 0 0 minecraft:warped_nylium",
                "fill 14 0 0 21 0 0 minecraft:cyan_wool",
                "fill 22 0 0 22 0 0 minecraft:b",
                "fill 23 0 0 36 0 0 minecraft:black_wool",
                "fill 0 0 1 9 0 1 minecraft:black_wool",
                "fill 10 0 1 10 0 1 minecraft:b",
                "fill 11 0 1 11 0 1 minecraft:cyan_wool",
                "fill 12 0 1 13 0 1 minecraft:light_blue_wool",
                "fill 14 0 1 23 0 1 minecraft:cyan_wool",
                "fill 24 0 1 24 0 1 minecraft:b",
                "fill 25 0 1 36 0 1 minecraft:black_wool",
                "fill 0 0 2 8 0 2 minecraft:black_wool",
                "fill 9 0 2 9 0 2 minecraft:gray_terracotta",
                "fill 10 0 2 10 0 2 minecraft:light_blue_wool",
                "fill 11 0 2 11 0 2 minecraft:cyan_wool",
                "fill 12 0 2 13 0 2 minecraft:warped_nylium",
                "fill 14 0 2 24 0 2 minecraft:cyan_wool",
                "fill 25 0 2 25 0 2 minecraft:b",
                "fill 26 0 2 36 0 2 minecraft:black_wool",
                "fill 0 0 3 8 0 3 minecraft:black_wool",
                "fill 9 0 3 9 0 3 minecraft:warped_nylium",
                "fill 10 0 3 10 0 3 minecraft:light_blue_wool",
                "fill 11 0 3 11 0 3 minecraft:b",
                "fill 12 0 3 13 0 3 minecraft:black_wool",
                "fill 14 0 3 14 0 3 minecraft:b",
                "fill 15 0 3 25 0 3 minecraft:cyan_wool",
                "fill 26 0 3 36 0 3 minecraft:black_wool",
                "fill 0 0 4 8 0 4 minecraft:black_wool",
                "fill 9 0 4 9 0 4 minecraft:cyan_wool",
                "fill 10 0 4 10 0 4 minecraft:light_blue_wool",
                "fill 11 0 4 11 0 4 minecraft:b",
                "fill 12 0 4 13 0 4 minecraft:black_wool",
                "fill 14 0 4 14 0 4 minecraft:b",
                "fill 15 0 4 25 0 4 minecraft:cyan_wool",
                "fill 26 0 4 26 0 4 minecraft:gray_terracotta",
                "fill 27 0 4 36 0 4 minecraft:black_wool",
                "fill 0 0 5 8 0 5 minecraft:black_wool",
                "fill 9 0 5 9 0 5 minecraft:cyan_wool",
                "fill 10 0 5 10 0 5 minecraft:light_blue_wool",
                "fill 11 0 5 11 0 5 minecraft:cyan_wool",
                "fill 12 0 5 13 0 5 minecraft:black_wool",
                "fill 14 0 5 25 0 5 minecraft:cyan_wool",
                "fill 26 0 5 26 0 5 minecraft:b",
                "fill 27 0 5 36 0 5 minecraft:black_wool",
                "fill 0 0 6 8 0 6 minecraft:black_wool",
                "fill 9 0 6 25 0 6 minecraft:cyan_wool",
                "fill 26 0 6 26 0 6 minecraft:gray_terracotta",
                "fill 27 0 6 36 0 6 minecraft:black_wool",
                "fill 0 0 7 8 0 7 minecraft:black_wool",
                "fill 9 0 7 9 0 7 minecraft:warped_nylium",
                "fill 10 0 7 25 0 7 minecraft:cyan_wool",
                "fill 26 0 7 26 0 7 minecraft:gray_terracotta",
                "fill 27 0 7 36 0 7 minecraft:black_wool",
                "fill 0 0 8 17 0 8 minecraft:black_wool",
                "fill 18 0 8 25 0 8 minecraft:cyan_wool",
                "fill 26 0 8 26 0 8 minecraft:gray_terracotta",
                "fill 27 0 8 36 0 8 minecraft:black_wool",
                "fill 0 0 9 2 0 9 minecraft:black_wool",
                "fill 3 0 9 3 0 9 minecraft:gray_terracotta",
                "fill 4 0 9 25 0 9 minecraft:cyan_wool",
                "fill 26 0 9 26 0 9 minecraft:black_wool",
                "fill 27 0 9 27 0 9 minecraft:dark_oak_planks",
                "fill 28 0 9 30 0 9 minecraft:gold_block",
                "fill 31 0 9 31 0 9 minecraft:r",
                "fill 32 0 9 32 0 9 minecraft:oak_planks",
                "fill 33 0 9 36 0 9 minecraft:black_wool",
                "fill 0 0 10 1 0 10 minecraft:black_wool",
                "fill 2 0 10 2 0 10 minecraft:warped_nylium",
                "fill 3 0 10 4 0 10 minecraft:light_blue_wool",
                "fill 5 0 10 25 0 10 minecraft:cyan_wool",
                "fill 26 0 10 26 0 10 minecraft:black_wool",
                "fill 27 0 10 27 0 10 minecraft:l",
                "fill 28 0 10 28 0 10 minecraft:sand",
                "fill 29 0 10 32 0 10 minecraft:gold_block",
                "fill 33 0 10 33 0 10 minecraft:y",
                "fill 34 0 10 36 0 10 minecraft:black_wool",
                "fill 0 0 11 0 0 11 minecraft:black_wool",
                "fill 1 0 11 1 0 11 minecraft:b",
                "fill 2 0 11 2 0 11 minecraft:light_blue_wool",
                "fill 3 0 11 25 0 11 minecraft:cyan_wool",
                "fill 26 0 11 26 0 11 minecraft:black_wool",
                "fill 27 0 11 27 0 11 minecraft:l",
                "fill 28 0 11 33 0 11 minecraft:gold_block",
                "fill 34 0 11 34 0 11 minecraft:l",
                "fill 35 0 11 36 0 11 minecraft:black_wool",
                "fill 0 0 12 0 0 12 minecraft:black_wool",
                "fill 1 0 12 25 0 12 minecraft:cyan_wool",
                "fill 26 0 12 26 0 12 minecraft:black_wool",
                "fill 27 0 12 27 0 12 minecraft:l",
                "fill 28 0 12 33 0 12 minecraft:gold_block",
                "fill 34 0 12 34 0 12 minecraft:sponge",
                "fill 35 0 12 36 0 12 minecraft:black_wool",
                "fill 0 0 13 0 0 13 minecraft:b",
                "fill 1 0 13 1 0 13 minecraft:light_blue_wool",
                "fill 2 0 13 25 0 13 minecraft:cyan_wool",
                "fill 26 0 13 26 0 13 minecraft:black_wool",
                "fill 27 0 13 27 0 13 minecraft:l",
                "fill 28 0 13 34 0 13 minecraft:gold_block",
                "fill 35 0 13 35 0 13 minecraft:b",
                "fill 36 0 13 36 0 13 minecraft:black_wool",
                "fill 0 0 14 0 0 14 minecraft:warped_nylium",
                "fill 1 0 14 1 0 14 minecraft:light_blue_wool",
                "fill 2 0 14 25 0 14 minecraft:cyan_wool",
                "fill 26 0 14 26 0 14 minecraft:black_wool",
                "fill 27 0 14 27 0 14 minecraft:y",
                "fill 28 0 14 34 0 14 minecraft:gold_block",
                "fill 35 0 14 35 0 14 minecraft:g",
                "fill 36 0 14 36 0 14 minecraft:black_wool",
                "fill 0 0 15 24 0 15 minecraft:cyan_wool",
                "fill 25 0 15 25 0 15 minecraft:b",
                "fill 26 0 15 26 0 15 minecraft:gray_terracotta",
                "fill 27 0 15 34 0 15 minecraft:gold_block",
                "fill 35 0 15 35 0 15 minecraft:oak_planks",
                "fill 36 0 15 36 0 15 minecraft:black_wool",
                "fill 0 0 16 22 0 16 minecraft:cyan_wool",
                "fill 23 0 16 23 0 16 minecraft:warped_nylium",
                "fill 24 0 16 24 0 16 minecraft:black_wool",
                "fill 25 0 16 25 0 16 minecraft:b",
                "fill 26 0 16 26 0 16 minecraft:sponge",
                "fill 27 0 16 34 0 16 minecraft:gold_block",
                "fill 35 0 16 35 0 16 minecraft:y",
                "fill 36 0 16 36 0 16 minecraft:black_wool",
                "fill 0 0 17 10 0 17 minecraft:cyan_wool",
                "fill 11 0 17 11 0 17 minecraft:b",
                "fill 12 0 17 23 0 17 minecraft:black_wool",
                "fill 24 0 17 24 0 17 minecraft:g",
                "fill 25 0 17 25 0 17 minecraft:sponge",
                "fill 26 0 17 34 0 17 minecraft:gold_block",
                "fill 35 0 17 35 0 17 minecraft:y",
                "fill 36 0 17 36 0 17 minecraft:black_wool",
                "fill 0 0 18 9 0 18 minecraft:cyan_wool",
                "fill 10 0 18 10 0 18 minecraft:black_wool",
                "fill 11 0 18 11 0 18 minecraft:gray_terracotta",
                "fill 12 0 18 12 0 18 minecraft:oak_planks",
                "fill 13 0 18 22 0 18 minecraft:y",
                "fill 23 0 18 23 0 18 minecraft:sponge",
                "fill 24 0 18 34 0 18 minecraft:gold_block",
                "fill 35 0 18 35 0 18 minecraft:y",
                "fill 36 0 18 36 0 18 minecraft:black_wool",
                "fill 0 0 19 8 0 19 minecraft:cyan_wool",
                "fill 9 0 19 9 0 19 minecraft:black_wool",
                "fill 10 0 19 10 0 19 minecraft:dark_oak_planks",
                "fill 11 0 19 34 0 19 minecraft:gold_block",
                "fill 35 0 19 35 0 19 minecraft:y",
                "fill 36 0 19 36 0 19 minecraft:black_wool",
                "fill 0 0 20 7 0 20 minecraft:cyan_wool",
                "fill 8 0 20 8 0 20 minecraft:b",
                "fill 9 0 20 9 0 20 minecraft:gray_terracotta",
                "fill 10 0 20 34 0 20 minecraft:gold_block",
                "fill 35 0 20 35 0 20 minecraft:podzol",
                "fill 36 0 20 36 0 20 minecraft:black_wool",
                "fill 0 0 21 0 0 21 minecraft:b",
                "fill 1 0 21 7 0 21 minecraft:cyan_wool",
                "fill 8 0 21 8 0 21 minecraft:black_wool",
                "fill 9 0 21 9 0 21 minecraft:oak_planks",
                "fill 10 0 21 34 0 21 minecraft:gold_block",
                "fill 35 0 21 35 0 21 minecraft:brown_terracotta",
                "fill 36 0 21 36 0 21 minecraft:black_wool",
                "fill 0 0 22 0 0 22 minecraft:black_wool",
                "fill 1 0 22 7 0 22 minecraft:cyan_wool",
                "fill 8 0 22 8 0 22 minecraft:black_wool",
                "fill 9 0 22 9 0 22 minecraft:acacia_planks",
                "fill 10 0 22 33 0 22 minecraft:gold_block",
                "fill 34 0 22 34 0 22 minecraft:sponge",
                "fill 35 0 22 36 0 22 minecraft:black_wool",
                "fill 0 0 23 0 0 23 minecraft:black_wool",
                "fill 1 0 23 7 0 23 minecraft:cyan_wool",
                "fill 8 0 23 8 0 23 minecraft:black_wool",
                "fill 9 0 23 9 0 23 minecraft:acacia_planks",
                "fill 10 0 23 33 0 23 minecraft:gold_block",
                "fill 34 0 23 34 0 23 minecraft:oak_planks",
                "fill 35 0 23 36 0 23 minecraft:black_wool",
                "fill 0 0 24 1 0 24 minecraft:black_wool",
                "fill 2 0 24 7 0 24 minecraft:cyan_wool",
                "fill 8 0 24 8 0 24 minecraft:black_wool",
                "fill 9 0 24 9 0 24 minecraft:acacia_planks",
                "fill 10 0 24 32 0 24 minecraft:gold_block",
                "fill 33 0 24 33 0 24 minecraft:sponge",
                "fill 34 0 24 34 0 24 minecraft:b",
                "fill 35 0 24 36 0 24 minecraft:black_wool",
                "fill 0 0 25 1 0 25 minecraft:black_wool",
                "fill 2 0 25 2 0 25 minecraft:b",
                "fill 3 0 25 7 0 25 minecraft:cyan_wool",
                "fill 8 0 25 8 0 25 minecraft:black_wool",
                "fill 9 0 25 9 0 25 minecraft:acacia_planks",
                "fill 10 0 25 31 0 25 minecraft:gold_block",
                "fill 32 0 25 32 0 25 minecraft:sponge",
                "fill 33 0 25 33 0 25 minecraft:brown_terracotta",
                "fill 34 0 25 36 0 25 minecraft:black_wool",
                "fill 0 0 26 8 0 26 minecraft:black_wool",
                "fill 9 0 26 9 0 26 minecraft:acacia_planks",
                "fill 10 0 26 16 0 26 minecraft:gold_block",
                "fill 17 0 26 17 0 26 minecraft:sponge",
                "fill 18 0 26 26 0 26 minecraft:g",
                "fill 27 0 26 30 0 26 minecraft:dark_oak_planks",
                "fill 31 0 26 31 0 26 minecraft:g",
                "fill 32 0 26 32 0 26 minecraft:b",
                "fill 33 0 26 36 0 26 minecraft:black_wool",
                "fill 0 0 27 8 0 27 minecraft:black_wool",
                "fill 9 0 27 9 0 27 minecraft:acacia_planks",
                "fill 10 0 27 16 0 27 minecraft:gold_block",
                "fill 17 0 27 17 0 27 minecraft:sponge",
                "fill 18 0 27 25 0 27 minecraft:podzol",
                "fill 26 0 27 26 0 27 minecraft:gray_terracotta",
                "fill 27 0 27 36 0 27 minecraft:black_wool",
                "fill 0 0 28 8 0 28 minecraft:black_wool",
                "fill 9 0 28 9 0 28 minecraft:acacia_planks",
                "fill 10 0 28 25 0 28 minecraft:gold_block",
                "fill 26 0 28 26 0 28 minecraft:podzol",
                "fill 27 0 28 36 0 28 minecraft:black_wool",
                "fill 0 0 29 8 0 29 minecraft:black_wool",
                "fill 9 0 29 9 0 29 minecraft:acacia_planks",
                "fill 10 0 29 20 0 29 minecraft:gold_block",
                "fill 21 0 29 21 0 29 minecraft:sponge",
                "fill 22 0 29 22 0 29 minecraft:y",
                "fill 23 0 29 23 0 29 minecraft:acacia_planks",
                "fill 24 0 29 25 0 29 minecraft:gold_block",
                "fill 26 0 29 26 0 29 minecraft:podzol",
                "fill 27 0 29 36 0 29 minecraft:black_wool",
                "fill 0 0 30 8 0 30 minecraft:black_wool",
                "fill 9 0 30 9 0 30 minecraft:acacia_planks",
                "fill 10 0 30 20 0 30 minecraft:gold_block",
                "fill 21 0 30 21 0 30 minecraft:g",
                "fill 22 0 30 22 0 30 minecraft:black_wool",
                "fill 23 0 30 23 0 30 minecraft:b",
                "fill 24 0 30 24 0 30 minecraft:acacia_planks",
                "fill 25 0 30 25 0 30 minecraft:gold_block",
                "fill 26 0 30 26 0 30 minecraft:podzol",
                "fill 27 0 30 36 0 30 minecraft:black_wool",
                "fill 0 0 31 8 0 31 minecraft:black_wool",
                "fill 9 0 31 9 0 31 minecraft:y",
                "fill 10 0 31 19 0 31 minecraft:gold_block",
                "fill 20 0 31 20 0 31 minecraft:sponge",
                "fill 21 0 31 21 0 31 minecraft:b",
                "fill 22 0 31 23 0 31 minecraft:black_wool",
                "fill 24 0 31 24 0 31 minecraft:y",
                "fill 25 0 31 25 0 31 minecraft:gold_block",
                "fill 26 0 31 26 0 31 minecraft:dark_oak_planks",
                "fill 27 0 31 36 0 31 minecraft:black_wool",
                "fill 0 0 32 8 0 32 minecraft:black_wool",
                "fill 9 0 32 9 0 32 minecraft:brown_terracotta",
                "fill 10 0 32 20 0 32 minecraft:gold_block",
                "fill 21 0 32 21 0 32 minecraft:y",
                "fill 22 0 32 22 0 32 minecraft:gray_terracotta",
                "fill 23 0 32 23 0 32 minecraft:dark_oak_planks",
                "fill 24 0 32 25 0 32 minecraft:sponge",
                "fill 26 0 32 26 0 32 minecraft:b",
                "fill 27 0 32 36 0 32 minecraft:black_wool",
                "fill 0 0 33 9 0 33 minecraft:black_wool",
                "fill 10 0 33 10 0 33 minecraft:podzol",
                "fill 11 0 33 23 0 33 minecraft:gold_block",
                "fill 24 0 33 24 0 33 minecraft:sponge",
                "fill 25 0 33 25 0 33 minecraft:brown_terracotta",
                "fill 26 0 33 36 0 33 minecraft:black_wool",
                "fill 0 0 34 10 0 34 minecraft:black_wool",
                "fill 11 0 34 11 0 34 minecraft:gray_terracotta",
                "fill 12 0 34 12 0 34 minecraft:y",
                "fill 13 0 34 13 0 34 minecraft:sponge",
                "fill 14 0 34 21 0 34 minecraft:gold_block",
                "fill 22 0 34 22 0 34 minecraft:sponge",
                "fill 23 0 34 23 0 34 minecraft:oak_planks",
                "fill 24 0 34 24 0 34 minecraft:b",
                "fill 25 0 34 36 0 34 minecraft:black_wool",
                "fill 0 0 35 12 0 35 minecraft:black_wool",
                "fill 13 0 35 13 0 35 minecraft:b",
                "fill 14 0 35 14 0 35 minecraft:brown_terracotta",
                "fill 15 0 35 15 0 35 minecraft:podzol",
                "fill 16 0 35 19 0 35 minecraft:y",
                "fill 20 0 35 20 0 35 minecraft:podzol",
                "fill 21 0 35 21 0 35 minecraft:gray_terracotta",
                "fill 22 0 35 36 0 35 minecraft:black_wool",
                "fill 0 0 36 36 0 36 minecraft:black_wool",
                "fill 0 0 37 36 0 37 minecraft:black_wool",
                "fill 0 0 38 12 0 38 minecraft:black_wool",
                "fill 13 0 38 13 0 38 minecraft:gray_terracotta",
                "fill 14 0 38 16 0 38 minecraft:acacia_wood",
                "fill 17 0 38 18 0 38 minecraft:c",
                "fill 19 0 38 21 0 38 minecraft:acacia_wood",
                "fill 22 0 38 22 0 38 minecraft:gray_terracotta",
                "fill 23 0 38 36 0 38 minecraft:black_wool",
                "fill 0 0 39 8 0 39 minecraft:black_wool",
                "fill 9 0 39 9 0 39 minecraft:acacia_wood",
                "fill 10 0 39 10 0 39 minecraft:stone",
                "fill 11 0 39 11 0 39 minecraft:l",
                "fill 12 0 39 15 0 39 minecraft:light_gray_wool",
                "fill 16 0 39 19 0 39 minecraft:iron_block",
                "fill 20 0 39 23 0 39 minecraft:light_gray_wool",
                "fill 24 0 39 24 0 39 minecraft:l",
                "fill 25 0 39 25 0 39 minecraft:d",
                "fill 26 0 39 26 0 39 minecraft:gray_terracotta",
                "fill 27 0 39 36 0 39 minecraft:black_wool",
            ],
            "mustnt_save",
            "1.18.2",
        )


ascii_expected_default = """@@@@@@@@@@@@@@@@@@@@@@@@@%%##*****+++++++******###%@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@@@#*+++++++++++++++++************#@@@@@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@%+++++####*++++++++****************#@@@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@*+++*@@@@@@#+++++********************@@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@++++*@@@@@@#+++**********************@@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@+++++*####*++************************@@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@+++++++++++**************************@@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@###################******************@@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@%%%%%%%%%%%%%%%%%%%%%%%%%%%%#******************@@%********#%@@@@@@@@@@
@@@@@%#*+++++++++++++++++********************************@@*:::::::::::-*@@@@@@@
@@@%+++++++++++++++++++**********************************@@*:::::::::::::-%@@@@@
@@*++++++++++++++++++************************************@@*:::::::::::::::%@@@@
@#+++++++++++++++++**************************************@@*:::::::::::::::-@@@@
@++++++++++++++++***************************************#@@-::::::::::::::::*@@@
#++++++++++++++****************************************%@@-:::::::::::::::::-@@@
*++++++++++++***************************************#%@@*::::::::::::::::::::%@@
+++++++++++*************#%%%@@@@@@@@@@@@@@@@@@@@@@@@%*=::::::::::::::::::::::#@@
+++++++++************#@@%*+=----------------------:::::::::::::::::::::::::::#@@
*++++++************#@@*-:::::::::::::::::::::::::::::::::::::::::::::::::::::%@@
%++++*************%@%-::::::::::::::::::::::::::::::::::::::::::::::::::::::=@@@
@*+***************@@-:::::::::::::::::::::::::::::::::::::::::::::::::::::::%@@@
@@***************#@@:::::::::::::::::::::::::::::::::::::::::::::::::::::::#@@@@
@@%**************#@@::::::::::::::::::::::::::::::::::::::::::::::::::::::*@@@@@
@@@@#************#@@::::::::::::::::::::::::::::::::::::::::::::::::::::=%@@@@@@
@@@@@@%##********#@@:::::::::::::::::::-----------------------------=+*%@@@@@@@@
@@@@@@@@@@@@@@@@@@@@::::::::::::::::::=@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@:::::::::::::::::::::::::::::::::::::#@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@:::::::::::::::::::::::::::::-:::::::#@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@::::::::::::::::::::::::::*@@@@#-::::#@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@-::::::::::::::::::::::::=@@@@@@*::::%@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@%-::::::::::::::::::::::::+%@@%*-:::*@@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@@@*-:::::::::::::::::::::::::::::-+%@@@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@@@@@@#+=-::::::::::::::::::::=+*%@@@@@@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@%##**+++++++**##%@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
@@@@@@@@@@@@@@@@@@@@%%###*****+++++++++++++++++*****###%%%@@@@@@@@@@@@@@@@@@@@@@"""

ascii_expected_30_cols = """@@@@@@@@#***+++****#%@@@@@@@@@
@@@@@@@#+%@%++*******%@@@@@@@@
@@@@@@@%*************%@@@@@@@@
@@#**********#*******%+---+%@@
%+++++++*************%=:::::#@
*+++++**************#*::::::-@
++++****#*+++++++++=:::::::::@
#+****##::::::::::::::::::::+@
@#****#+:::::::::::::::::::=@@
@@@%%%@+::::::-++++++*####%@@@
@@@@@@@*:::::::::=*=:*@@@@@@@@
@@@@@@@%-::::::::+#==%@@@@@@@@
@@@@@@@@@@#**+++**#@@@@@@@@@@@
@@@@@@@@%%#########%%@@@@@@@@@"""


class TestToASCII(unittest.TestCase):
    def test_default(self):
        self.assertEqual(to_ascii(image=IMG_PATH), ascii_expected_default)

    def test_custom_cols(self):
        self.assertEqual(to_ascii(image=IMG_PATH, cols=30), ascii_expected_30_cols)


if __name__ == "__main__":
    unittest.main()

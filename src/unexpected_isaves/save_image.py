import json
import os
from contextlib import suppress
from math import sqrt
from typing import Optional, Tuple, Union

import numpy as np
import pandas as pd
from openpyxl import load_workbook, styles, utils
from PIL import Image


def to_excel(
    image: Union[Image.Image, str],
    path: str,
    lower_image_size_by: int = 10,
    image_position: Tuple[int, int] = (0, 0),
    **spreadsheet_kwargs,
) -> None:
    """
    - Added on release 0.0.1;
    - Coded originally on https://github.com/Eric-Mendes/image2excel

    Saves an image as a `.xlsx` file by coloring its cells each pixel's color.

    ## Parameters
    * :param image: Your image opened using the `PIL.Image` module or the image's path as `str`;
    * :param path: The path that you want to save your output file.
    Example: `/home/user/Documents/my_image.xlsx`;
    * :param lower_image_size_by: A factor that the function will divide
    your image's dimensions by. Defaults to `10`;
       * It is very important that you lower your image's dimensions because a big image might take the function a long time to process plus your spreadsheet will probably take a long time to load on any software that you use to open it;
    * :param image_position: a tuple determining the position of the top leftmost pixel. Cannot have negative values. Defaults to `(0,0)`.
    * :param **spreadsheet_kwargs: See below.

    ## Spreadsheet Kwargs
    Optional parameters to tweak the spreadsheet's appearance.

    * :param row_height (`float`): the rows' height. Defaults to `15`;
    * :param column_width (`float`): the columns' width. Defaults to `2.3`;
       * The default values on `row_height` and `column_width` were specifically thought out so that they make the cells squared, however - as any hardcoded value - they might not do the trick on your device. That is when you might want to tweak them a little bit.
    * :param delete_cell_value (`bool`): wheter to keep or not the text corresponding to that color. Defaults to `True`;
    * :param zoom_scale (`int`): how much to zoom in or out on the spreadsheet. Defaults to `20` which seems to be the default max zoom out on most spreadsheet softwares.

    ## Return
    * :return: `None`, but outputs a `.xlsx` file on the given `path`.
    """
    if isinstance(image, str):
        image = Image.open(image)

    image_position_row = image_position[0]
    image_position_col = image_position[1]
    if image_position_row > 0:
        image_position_row -= 1
    if image_position_col > 0:
        image_position_col -= 1
    if image_position_row < 0 or image_position_col < 0:
        raise ValueError("image_position cannot have negative values.")

    image = image.convert("RGB")
    image = image.resize(
        (image.size[0] // lower_image_size_by, image.size[1] // lower_image_size_by)
    )
    # OpenPyxl colors work in a weird way
    image_colors_processed = [
        ["%02x%02x%02x" % tuple(item) for item in row]
        for row in np.array(image).tolist()
    ]

    df = pd.DataFrame(image_colors_processed)
    image_name = os.path.splitext(os.path.split(path)[1])[0]

    # Saving a DataFrame where each cell has a text corresponding to the RGB color its background should be
    df.to_excel(
        path,
        index=False,
        header=False,
        startrow=image_position_row,
        startcol=image_position_col,
    )

    # Loading the excel file, painting each cell with its color and saving the updates
    wb = load_workbook(path)

    ws = wb.active
    ws.title = image_name

    for row in range(1, df.shape[0] + 1):
        for col in range(1, df.shape[1] + 1):
            cell = ws.cell(
                row=row + image_position_row, column=col + image_position_col
            )
            # Makes cells squared
            ws.row_dimensions[row + image_position_row].height = spreadsheet_kwargs.get(
                "row_height", 15
            )
            ws.column_dimensions[
                utils.get_column_letter(col + image_position_col)
            ].width = spreadsheet_kwargs.get("column_width", 2.3)

            # Painting the cell
            cell.fill = styles.PatternFill(
                start_color=cell.value, end_color=cell.value, fill_type="solid"
            )
            if spreadsheet_kwargs.get("delete_cell_value", True):
                cell.value = None  # Deletes the text from the cell

    # Saves spreadsheet already zoomed in or out
    ws.sheet_view.zoomScale = spreadsheet_kwargs.get("zoom_scale", 20)
    wb.save(path)


def to_minecraft(
    image: Union[Image.Image, str],
    path: str,
    lower_image_size_by: int = 10,
    player_pos: Tuple[int, int, int] = (0, 0, 0),
    minecraft_version: str = "1.18.2",
) -> None:
    """
    - Added on release 0.0.1;
    - Coded originally on https://github.com/Eric-Mendes/pixel-art-map

    Saves an image as a minecraft datapack that when loaded into your world will build a pixel art of it on the player's position.

    ## Parameters
    * :param image: Your image opened using the `PIL.Image` module or the image's path as `str`;
    * :param path: The path that you want to save your datapack.
    Example: `/home/user/Documents/my_image_datapack`;
    * :param lower_image_size_by: A factor that the function will divide
    your image's dimensions by. Defaults to `10`;
    * :param player_pos: The player's (x, y, z) position. Defaults to `(0, 0, 0)`.
    * :param minecraft_version: The minecraft version. Needs to be higher than or equal to `1.13.0`, and defaults to `1.19.0`.

    ## Return
    * :return: `None`, but outputs a datapack on the given `path`.
    """
    if isinstance(image, str):
        image = Image.open(image)

    image = image.convert("RGB")

    # Makes the commands that the datapack will run when loaded
    def script(df, **kwargs):
        player_pos = [
            kwargs.get("player_x", 0),
            kwargs.get("player_y", 0),
            kwargs.get("player_z", 0),
        ]
        z = (df != df.shift()).cumsum()
        zri = z.reset_index()
        ix_name = z.index.name
        co_name = z.columns.name
        for i in z:
            v = zri.groupby(i)[ix_name].agg(["first", "last"])
            s = {co_name: i}
            e = {co_name: i}
            for _, r in v.iterrows():
                s[ix_name] = r["first"]
                e[ix_name] = r["last"]
                material = df.loc[r["first"], i]
                yield f'fill {s["x"] + player_pos[0]} {0 + player_pos[1]} {s["z"] + player_pos[2]} {e["x"] + player_pos[0]} {0 + player_pos[1]} {e["z"] + player_pos[2]} {material.split(",")[0].strip()}'

    # Helper function. Loads the blocks an the colors they have when looked at via map,
    # and maps the pixels to the blocks
    blocks = json.load("blocks.json")

    def to_minecraft_color(pxl):
        color = None
        min_distance = None
        for item in blocks:
            # Calculates the "distance" between two RGB colors as if they
            # were points in a 3-dimensional space.
            # The closer they are, the more they look like each other.
            euclidean_distance = sqrt(
                sum([pow(p - c, 2) for p, c in zip(item["rgb"], pxl)])
            )

            if min_distance is None or euclidean_distance < min_distance:
                min_distance = euclidean_distance
                color = ", ".join("minecraft:" + block for block in item["blocks"])
        return color

    # Resizing the image and mapping each pixel's color to a minecraft color
    image = image.resize(
        (image.size[0] // lower_image_size_by, image.size[1] // lower_image_size_by)
    )
    image_colors_processed = [
        [to_minecraft_color(pixel) for pixel in row] for row in np.array(image)
    ]

    # Getting the name that the image should have via the given path
    image_name = os.path.splitext(os.path.split(path)[1])[0]

    df = pd.DataFrame(image_colors_processed)

    # Creates - in an error proof manner - the folder structure of the datapack
    with suppress(FileExistsError):
        os.makedirs(f"{path}/data/minecraft/tags/functions")
        os.makedirs(f"{path}/data/pixelart-map/functions")

    if minecraft_version >= "1.13.0" and minecraft_version <= "1.14.4":
        datapack_version = 4
    elif minecraft_version >= "1.15.0" and minecraft_version <= "1.16.1":
        datapack_version = 5
    elif minecraft_version >= "1.16.2" and minecraft_version <= "1.16.5":
        datapack_version = 6
    elif minecraft_version >= "1.17.0" and minecraft_version <= "1.17.1":
        datapack_version = 7
    elif minecraft_version >= "1.18.0" and minecraft_version <= "1.18.1":
        datapack_version = 8
    elif minecraft_version == "1.18.2":
        datapack_version = 9
    elif minecraft_version == "1.19.0":
        datapack_version = 10
    else:
        raise ValueError(
            "Unsupported minecraft_version. If you feel like this is a mistake, open an issue at https://github.com/Eric-Mendes/unexpected-isaves/issues to let us know."
        )

    pack_mcmeta = {
        "pack": {
            "pack_format": datapack_version,
            "description": f"This datapack will generate the image ({image_name}) in your world",
        }
    }
    load_json = {"values": ["pixelart-map:load"]}
    tick_json = {"values": ["pixelart-map:tick"]}

    with open(f"{path}/pack.mcmeta", "w") as file:
        file.write(json.dumps(pack_mcmeta, indent=4))
    with open(f"{path}/data/minecraft/tags/functions/load.json", "w") as file:
        file.write(json.dumps(load_json, indent=4))
    with open(f"{path}/data/minecraft/tags/functions/tick.json", "w") as file:
        file.write(json.dumps(tick_json, indent=4))

    with open(f"{path}/data/pixelart-map/functions/tick.mcfunction", "w") as file:
        file.write("")

    # Making the commands that when ran will build the image's pixel art.
    # This part's had a huge contribution from this thread: https://stackoverflow.com/questions/70512775/how-to-group-elements-in-dataframe-by-row/70546452#70546452
    df = df.rename_axis(index="z", columns="x")

    a = list(
        script(
            df,
            player_x=player_pos[0],
            player_y=player_pos[1],
            player_z=player_pos[2],
        )
    )
    b = list(
        script(
            df.T,
            player_x=player_pos[0],
            player_y=player_pos[1],
            player_z=player_pos[2],
        )
    )
    res = min([a, b], key=len)
    with open(f"{path}/data/pixelart-map/functions/load.mcfunction", "w") as file:
        file.write("\n".join(res))


def to_ascii(
    image: Union[Image.Image, str],
    path: Optional[str] = None,
    cols: int = 80,
    scale: float = 0.43,
    more_levels: bool = False,
) -> str:
    """
    Creates an ascii art out of an image.
    Credits - https://www.geeksforgeeks.org/converting-image-ascii-image-python/

    Args:
        image: Your image opened using the `PIL.Image` module or the image's path as `str`.
        path: The path that you want to save your `.txt` file, if you want to save it. Otherwise the function will only return the ascii art string.
        cols: Used for computing tile width. Defaults to `80`.
        scale: Used for computing tile height. Defaults to `0.43` - ok for a monospaced font like Courier.
        more_levels: When set to `True` uses more ascii characters (70). Defaults to `False` (10 ascii characters).

    Returns:
        The ascii art of the `image`.

    Raises:
        ValueError: "Image too small for specified cols."
    """
    # 70 levels of gray
    gscale1 = "$@B%8&WM#*oahkbdpqwmZO0QLCJUYXzcvunxrjft/\|()1{}[]?-_+~<>i!lI;:,\"^`'. "

    # 10 levels of gray
    gscale2 = "@%#*+=-:. "

    def getAverageL(image):
        """
        Given PIL Image, return average value of grayscale value
        """
        # get image as numpy array
        im = np.array(image)

        # get shape
        w, h = im.shape

        # get average
        return np.average(im.reshape(w * h))

    # open image and convert to grayscale
    if isinstance(image, str):
        image = Image.open(image)
    image = image.convert("L")

    # store dimensions
    W, H = image.size[0], image.size[1]

    # compute width of tile
    w = W / cols

    # compute tile height based on aspect ratio and scale
    h = w / scale

    # compute number of rows
    rows = int(H / h)

    # check if image size is too small
    if cols > W or rows > H:
        raise ValueError("Image too small for specified cols.")

    # ascii image is a list of character strings
    aimg = []
    # generate list of dimensions
    for j in range(rows):
        y1 = int(j * h)
        y2 = int((j + 1) * h)

        # correct last tile
        if j == rows - 1:
            y2 = H

        # append an empty string
        aimg.append("")

        for i in range(cols):

            # crop image to tile
            x1 = int(i * w)
            x2 = int((i + 1) * w)

            # correct last tile
            if i == cols - 1:
                x2 = W

            # crop image to extract tile
            img = image.crop((x1, y1, x2, y2))

            # get average luminance
            avg = int(getAverageL(img))

            # look up ascii char
            if more_levels:
                gsval = gscale1[int((avg * 69) / 255)]
            else:
                gsval = gscale2[int((avg * 9) / 255)]

            # append ascii char to string
            aimg[j] += gsval

    if path is not None:
        f = open(path, "w")

        # write to file
        for row in aimg:
            f.write(row + "\n")

        # cleanup
        f.close()

    # return txt image
    return "\n".join(aimg)

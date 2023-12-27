import os
from typing import List, Tuple, Union

import numpy as np
from openpyxl import Workbook, styles, utils
from PIL import Image


def _save(
    processed_pil_image: List[List[str]],
    path: Union[os.PathLike, str],
    image_position: Tuple[int, int],
    **spreadsheet_kwargs,
) -> None:
    starting_row, starting_col = image_position
    ending_row, ending_col = (
        starting_row + len(processed_pil_image),
        starting_col + len(processed_pil_image[0]),
    )

    wb = Workbook()
    ws = wb.active

    image_name = os.path.splitext(os.path.split(path)[1])[0]
    ws.title = image_name

    for row in range(starting_row, ending_row + 1):
        for col in range(starting_col, ending_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = processed_pil_image[row - starting_row - 1][
                col - starting_col - 1
            ]

            # Makes cells squared
            ws.row_dimensions[row].height = spreadsheet_kwargs.get("row_height", 15)
            ws.column_dimensions[
                utils.get_column_letter(col)
            ].width = spreadsheet_kwargs.get("column_width", 2.3)

            # Painting the cell
            cell.fill = styles.PatternFill(
                start_color=cell.value, end_color=cell.value, fill_type="solid"
            )
            if spreadsheet_kwargs.get("delete_cell_value", True):
                cell.value = None  # Deletes the text from the cell

    # Saves spreadsheet already zoomed in or out
    ws.sheet_view.zoomScale = spreadsheet_kwargs.get("zoom_scale", 20)
    wb.save(path)
    return None


def _load_image(image: Union[Image.Image, os.PathLike, str]) -> Image.Image:
    if isinstance(image, (os.PathLike, str)):
        if not os.path.exists(image):
            raise ValueError("Error loading image. Image path not found.")
        image = Image.open(image)
    return image


def _to_openpyxl_colors(image: Image.Image) -> List[List[str]]:
    image_colors_processed = [
        ["%02x%02x%02x" % tuple(item) for item in row]
        for row in np.array(image).tolist()
    ]
    return image_colors_processed


def _process(image: Image.Image, lower_image_size_by: int):
    image_rgb = image.convert("RGB")
    image_rgb_resized = image_rgb.resize(
        (
            image_rgb.size[0] // lower_image_size_by,
            image_rgb.size[1] // lower_image_size_by,
        )
    )
    image_openpyxl_colors_resized = _to_openpyxl_colors(image_rgb_resized)

    return image_openpyxl_colors_resized


def to_excel(
    image: Union[Image.Image, os.PathLike, str],
    path: Union[os.PathLike, str],
    lower_image_size_by: int = 10,
    image_position: Tuple[int, int] = (0, 0),
    **spreadsheet_kwargs,
) -> None:
    """
    - Coded originally on https://github.com/Eric-Mendes/image2excel

    Saves an image as a `.xlsx` file by coloring its cells each pixel's color.

    Args
        image: Your image opened using the `PIL.Image` module or the image's path.
        path: The path that you want to save your output file. Example: `/home/user/Documents/my_image.xlsx`.
        lower_image_size_by: A factor that the function will divide your image's dimensions by. Defaults to `10`. It is very important that you lower your image's dimensions because a big image might take the function a long time to process plus your spreadsheet will probably take a long time to load on any software that you use to open it.
        image_position: a tuple determining the position of the top leftmost pixel. Cannot have negative values. Defaults to `(0,0)`.
        **spreadsheet_kwargs: Optional parameters to tweak the spreadsheet's appearance. The default values on `row_height` and `column_width` were specifically thought out so that they make the cells squared, however - as any hardcoded value - they might not do the trick on your device. That is when you might want to tweak them a little bit.
            row_height (`float`): the rows' height. Defaults to `15`.
            column_width (`float`): the columns' width. Defaults to `2.3`.
            delete_cell_value (`bool`): wheter to keep or not the text corresponding to that color. Defaults to `True`.
            zoom_scale (`int`): how much to zoom in or out on the spreadsheet. Defaults to `20` which seems to be the default max zoom out on most spreadsheet softwares.

    Returns
        `None`, but outputs a `.xlsx` file on the given `path`.
    """
    if os.path.exists(path):
        raise ValueError(
            f"{path} already exists. Please provide a new path for your .xlsx."
        )

    if image_position[0] < 0 or image_position[1] < 0:
        raise ValueError("image_position cannot have negative values.")

    pil_image = _load_image(image)
    processed_pil_image = _process(pil_image, lower_image_size_by)
    image_position_processed = (
        image_position[0] + int(image_position[0] == 0),
        image_position[1] + int(image_position[1] == 0),
    )
    save_result = _save(
        processed_pil_image,
        path=path,
        image_position=image_position_processed,
        **spreadsheet_kwargs,
    )

    return save_result
